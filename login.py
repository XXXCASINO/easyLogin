from selenium import webdriver
import openpyxl


from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from webdriver_manager.chrome import ChromeDriverManager

import time

#################################################################################

customService = Service(ChromeDriverManager().install())
customOption = Options()

browser = webdriver.Chrome(service = customService, options = customOption)


URL = 'https://www.google.com/'
browser.get(URL)
browser.implicitly_wait(10)

#################################################################################

browser.find_element(By.XPATH, '//*[@id="L2AGLb"]/div').click
time.sleep(5)

busca = browser.find_element(By.XPATH, '//*[@id="APjFqb"]').send_keys("네이버 주식")
time.sleep(10)
busca = browser.find_element(By.XPATH, '//*[@id="APjFqb"]').click


#메일 값 획득
temp = browser.find_element(By.XPATH, '//*[@id="knowledge-finance-wholepage__entity-summary"]/div[3]/g-card-section/div/g-card-section/div[2]/div[1]/span[1]/span/span[1]').text
print(temp)


#send_keys
browser.find_element(By.XPATH, '//*[@id="L2AGLb"]/div').click()
time.sleep(1)



#엑셀파일 생성
xlsxFile = openpyxl.Workbook()

#생성한 파일에서 시트 생성
xlsxSheet = xlsxFile.active


#시트 특정 셀에 데이터 입력
for i in range(10):
	xlsxSheet.cell(row = 1, column = i + 1).value = browser.find_element
	#find_element().text 로 찾은 값을 넣으면 됨


#저장
xlsxFile.save('EXTA AUTO1.xlsx')



-------------------------------



from selenium import webdriver
import openpyxl


from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options

from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys

from webdriver_manager.chrome import ChromeDriverManager

import time

#################################################################################

customService = Service(ChromeDriverManager().install())
customOption = Options()

browser = webdriver.Chrome(service = customService, options = customOption)


URL = 'https://www.accounts-bc.com/signin'
browser.get(URL)
browser.implicitly_wait(10)

#################################################################################


#메일 값 획득
temp = browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[2]/div[1]/div/input').text
print(temp)


#만약 조금 상위 레벨에서 진행
temp = browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[2]/div[1]/div/input').text
print(temp)


#send_keys
browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[2]/div[1]/div/input').send_keys('juyeonglee911029@gmail.com')
time.sleep(1)

browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[2]/div[2]/div/div/input').send_keys('12Qwaszx!@')


#버튼 클릭
browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[3]/button').click()
time.sleep(1)


browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[2]/div/div[1]/div/input').send_keys('441224')
time.sleep(1)

#버튼 클릭
browser.find_element(By.XPATH, '//*[@id="root"]/ul/li[2]/div/ul/li[3]').click()
time.sleep(20)


#엑셀파일 생성
xlsxFile = openpyxl.Workbook()

#생성한 파일에서 시트 생성
xlsxSheet = xlsxFile.active


#시트 특정 셀에 데이터 입력
for i in range(10):
	xlsxSheet.cell(row = 1, column = i + 1).value = i + 1
	#find_element().text 로 찾은 값을 넣으면 됨


#저장
xlsxFile.save('EXTA AUTO.xlsx')




